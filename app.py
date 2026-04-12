
from flask import Flask, jsonify, render_template, request
from openpyxl import load_workbook
from pathlib import Path
from functools import lru_cache
import json, os
import requests, math
from datetime import datetime, date
import re

BASE_DIR = Path(__file__).resolve().parent
app = Flask(__name__, static_folder="static", template_folder="templates")
CONFIG_PATH = BASE_DIR / "model_config.xlsx"
PARK_DIMENSIONS_PATH = BASE_DIR / "static" / "data" / "park_dimensions.json"

MLB_BASE = "https://statsapi.mlb.com/api/v1"
MLB_LIVE_BASE = "https://statsapi.mlb.com/api/v1.1"
ODDS_TRADER_WEATHER_URL = "https://www.oddstrader.com/mlb/weather/"
BALLPARK_PAL_FACTORS_URL = "https://www.ballparkpal.com/ParkFactorsGet.php"
USER_AGENT = "MLBWeatherProjector/1.0"
ODDS_TRADER_ABBR_MAP = {
    "ARI": "AZ",
}

def api_get(url, params=None):
    r = requests.get(url, params=params, timeout=25, headers={"User-Agent": USER_AGENT})
    r.raise_for_status()
    return r.json()

@lru_cache(maxsize=1)
def load_config():
    wb = load_workbook(CONFIG_PATH, data_only=True)
    bp = wb["Ballparks"]
    park_dimensions = {}
    if PARK_DIMENSIONS_PATH.exists():
        park_dimensions = json.loads(PARK_DIMENSIONS_PATH.read_text(encoding="utf-8"))
    parks = {}
    for row in bp.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        team_id, team, abbr, venue, lat, lon, altitude_ft, park_factor, roof_type, logo_path = row
        official_logo = BASE_DIR / "static" / "logos-official" / f"{abbr}.png"
        outline_meta = park_dimensions.get(str(abbr), {})
        parks[int(team_id)] = {
            "team": team, "abbr": abbr, "venue": venue, "lat": lat, "lon": lon,
            "altitude_ft": float(altitude_ft), "park_factor": float(park_factor),
            "roof_type": roof_type,
            "logo_path": f"/static/logos-official/{abbr}.png" if official_logo.exists() else "/" + str(logo_path).replace("\\", "/"),
            "field_size_text": outline_meta.get("field_size_text"),
            "outfield_points": outline_meta.get("outfield_points", []),
            "outline_source_url": outline_meta.get("source_url"),
        }
    mw = wb["ModelWeights"]
    weights = {}
    for row in mw.iter_rows(min_row=2, values_only=True):
        if row[0]:
            weights[str(row[0])] = float(row[1])
    return {"parks": parks, "weights": weights}

def safe_div(a, b, fallback=0.0):
    return a / b if b else fallback

def stat_float(value, fallback=0.0):
    try:
        if value in (None, "", ".---", "-.--"):
            return fallback
        return float(value)
    except (TypeError, ValueError):
        return fallback

def normalize_degrees(value, fallback=0.0):
    try:
        return float(value) % 360
    except (TypeError, ValueError):
        return fallback

def signed_angle_diff(target_deg, origin_deg):
    return ((target_deg - origin_deg + 540) % 360) - 180

def direction_to_degrees(label):
    directions = {
        "N": 0.0, "NNE": 22.5, "NE": 45.0, "ENE": 67.5,
        "E": 90.0, "ESE": 112.5, "SE": 135.0, "SSE": 157.5,
        "S": 180.0, "SSW": 202.5, "SW": 225.0, "WSW": 247.5,
        "W": 270.0, "WNW": 292.5, "NW": 315.0, "NNW": 337.5,
    }
    return directions.get(str(label or "").upper())

def odds_trader_abbr(abbr):
    return ODDS_TRADER_ABBR_MAP.get(str(abbr or "").upper(), str(abbr or "").upper())

def default_park_factors():
    return {
        "game_pk": None,
        "runs": 0.0,
        "hr": 0.0,
        "doubles_triples": 0.0,
        "singles": 0.0,
        "runs_formatted": "0%",
        "hr_formatted": "0%",
        "doubles_triples_formatted": "0%",
        "singles_formatted": "0%",
        "venue": None,
        "game_time": None,
        "source": "ballparkpal-unavailable",
    }

def innings_to_outs(ip_value):
    if ip_value in (None, ""):
        return 0
    text = str(ip_value)
    if "." not in text:
        return int(text) * 3
    whole, frac = text.split(".", 1)
    return int(whole) * 3 + int(frac or 0)

def outs_to_ip(outs):
    whole = outs // 3
    remainder = outs % 3
    return float(f"{whole}.{remainder}")

def season_candidates(game_date, max_seasons=2):
    year = int(str(game_date)[:4])
    return [str(year - offset) for offset in range(max_seasons)]

def game_date_only(game_date):
    return str(game_date)[:10]

@lru_cache(maxsize=8192)
def get_player_stat_splits(person_id, group, stat_type, season=None, opposing_player_id=None):
    params = {"stats": stat_type, "group": group}
    if season:
        params["season"] = str(season)
    if opposing_player_id:
        params["opposingPlayerId"] = int(opposing_player_id)
    data = api_get(f"{MLB_BASE}/people/{person_id}/stats", params).get("stats", [])
    return (data[0].get("splits", []) if data else [])

def collect_recent_player_games(person_id, group, game_date, limit=4):
    cutoff = game_date_only(game_date)
    games = []
    for season in season_candidates(game_date, max_seasons=2):
        for split in get_player_stat_splits(person_id, group, "gameLog", season=season):
            if split.get("gameType") != "R":
                continue
            if split.get("date") and split["date"] < cutoff:
                games.append(split)
    games.sort(key=lambda x: x.get("date", ""), reverse=True)
    return games[:limit]

def summarize_recent_hitting(person_id, game_date, limit=4):
    games = collect_recent_player_games(person_id, "hitting", game_date, limit=limit)
    ab = sum(stat_float(g.get("stat", {}).get("atBats", 0)) for g in games)
    pa = sum(stat_float(g.get("stat", {}).get("plateAppearances", 0)) for g in games)
    hits = sum(stat_float(g.get("stat", {}).get("hits", 0)) for g in games)
    walks = sum(stat_float(g.get("stat", {}).get("baseOnBalls", 0)) for g in games)
    hbp = sum(stat_float(g.get("stat", {}).get("hitByPitch", 0)) for g in games)
    sf = sum(stat_float(g.get("stat", {}).get("sacFlies", 0)) for g in games)
    tb = sum(stat_float(g.get("stat", {}).get("totalBases", 0)) for g in games)
    hr = sum(stat_float(g.get("stat", {}).get("homeRuns", 0)) for g in games)
    runs = sum(stat_float(g.get("stat", {}).get("runs", 0)) for g in games)
    rbi = sum(stat_float(g.get("stat", {}).get("rbi", 0)) for g in games)
    avg = safe_div(hits, ab, 0.0)
    obp = safe_div(hits + walks + hbp, ab + walks + hbp + sf, 0.0)
    slg = safe_div(tb, ab, 0.0)
    return {
        "games": len(games),
        "avg": round(avg, 3),
        "obp": round(obp, 3),
        "slg": round(slg, 3),
        "ops": round(obp + slg, 3),
        "hr": int(hr),
        "runs": int(runs),
        "rbi": int(rbi),
        "pa": round(pa, 1),
        "dates": [g.get("date") for g in games],
    }

def summarize_recent_pitching(person_id, game_date, limit=4):
    games = collect_recent_player_games(person_id, "pitching", game_date, limit=limit)
    outs = sum(innings_to_outs(g.get("stat", {}).get("inningsPitched", 0)) for g in games)
    ip = outs / 3 if outs else 0.0
    er = sum(stat_float(g.get("stat", {}).get("earnedRuns", 0)) for g in games)
    hits = sum(stat_float(g.get("stat", {}).get("hits", 0)) for g in games)
    walks = sum(stat_float(g.get("stat", {}).get("baseOnBalls", 0)) for g in games)
    ks = sum(stat_float(g.get("stat", {}).get("strikeOuts", 0)) for g in games)
    starts = sum(stat_float(g.get("stat", {}).get("gamesStarted", 0)) for g in games)
    return {
        "games": len(games),
        "era": round(safe_div(er * 9, ip, 4.20), 2),
        "whip": round(safe_div(hits + walks, ip, 1.30), 2),
        "k9": round(safe_div(ks * 9, ip, 8.5), 2),
        "ip_per_start": round(safe_div(ip, starts or len(games), 5.4), 2),
        "dates": [g.get("date") for g in games],
    }

def aggregate_vs_player_history(person_id, group, opposing_player_id):
    splits = get_player_stat_splits(person_id, group, "vsPlayer", opposing_player_id=opposing_player_id)
    ab = sum(stat_float(s.get("stat", {}).get("atBats", 0)) for s in splits)
    pa = sum(stat_float(s.get("stat", {}).get("plateAppearances", 0)) for s in splits)
    hits = sum(stat_float(s.get("stat", {}).get("hits", 0)) for s in splits)
    walks = sum(stat_float(s.get("stat", {}).get("baseOnBalls", 0)) for s in splits)
    hbp = sum(stat_float(s.get("stat", {}).get("hitByPitch", 0)) for s in splits)
    sf = sum(stat_float(s.get("stat", {}).get("sacFlies", 0)) for s in splits)
    tb = sum(stat_float(s.get("stat", {}).get("totalBases", 0)) for s in splits)
    hr = sum(stat_float(s.get("stat", {}).get("homeRuns", 0)) for s in splits)
    ks = sum(stat_float(s.get("stat", {}).get("strikeOuts", 0)) for s in splits)
    games = sum(stat_float(s.get("stat", {}).get("gamesPlayed", 0)) for s in splits)
    ip = sum(innings_to_outs(s.get("stat", {}).get("inningsPitched", 0)) for s in splits) / 3
    avg = safe_div(hits, ab, 0.0)
    obp = safe_div(hits + walks + hbp, ab + walks + hbp + sf, 0.0)
    slg = safe_div(tb, ab, 0.0)
    return {
        "games": int(games),
        "pa": round(pa, 1),
        "ab": int(ab),
        "hits": int(hits),
        "hr": int(hr),
        "strikeouts": int(ks),
        "avg": round(avg, 3),
        "obp": round(obp, 3),
        "slg": round(slg, 3),
        "ops": round(obp + slg, 3),
        "era_allowed": round(safe_div(sum(stat_float(s.get("stat", {}).get("earnedRuns", 0)) for s in splits) * 9, ip, 4.20), 2) if ip else None,
    }

@lru_cache(maxsize=1024)
def get_recent_team_games(team_id, game_date, limit=8):
    cutoff = game_date_only(game_date)
    games = []
    for year in season_candidates(game_date, max_seasons=2):
        start_date = f"{year}-01-01"
        end_date = cutoff if year == cutoff[:4] else f"{year}-12-31"
        data = api_get(
            f"{MLB_BASE}/schedule",
            {"sportId": 1, "teamId": int(team_id), "startDate": start_date, "endDate": end_date, "hydrate": "linescore"},
        )
        for day in data.get("dates", []):
            for game in day.get("games", []):
                if game.get("gameType") != "R":
                    continue
                if game.get("status", {}).get("codedGameState") != "F":
                    continue
                if game.get("officialDate") and game["officialDate"] >= cutoff:
                    continue
                away = game.get("teams", {}).get("away", {})
                home = game.get("teams", {}).get("home", {})
                if away.get("team", {}).get("id") == int(team_id):
                    scored = stat_float(away.get("score", 0))
                    allowed = stat_float(home.get("score", 0))
                    opp_abbr = home.get("team", {}).get("name", "UNK")[:3].upper()
                    desc = f"@{opp_abbr} {int(scored)}-{int(allowed)}"
                elif home.get("team", {}).get("id") == int(team_id):
                    scored = stat_float(home.get("score", 0))
                    allowed = stat_float(away.get("score", 0))
                    opp_abbr = away.get("team", {}).get("name", "UNK")[:3].upper()
                    desc = f"vs{opp_abbr} {int(scored)}-{int(allowed)}"
                else:
                    continue
                date_fmt = game.get("officialDate", "")[5:].replace("-", "/")
                games.append({"date": game.get("officialDate"), "desc": f"{date_fmt} {desc}", "runs_scored": scored, "runs_allowed": allowed})
    games.sort(key=lambda x: x["date"], reverse=True)
    return games[:limit]

def summarize_recent_team_runs(team_id, game_date, limit=4):
    games = get_recent_team_games(team_id, game_date, limit=limit)
    return {
        "games": len(games),
        "runs_per_game": round(safe_div(sum(g["runs_scored"] for g in games), len(games), 4.45), 2),
        "runs_allowed_per_game": round(safe_div(sum(g["runs_allowed"] for g in games), len(games), 4.45), 2),
        "dates": [g.get("desc", g["date"]) for g in games],
    }

@lru_cache(maxsize=512)
def summarize_park_recent_scoring(venue_id, game_date, limit=8):
    cutoff = game_date_only(game_date)
    games = []
    for year in season_candidates(game_date, max_seasons=2):
        start_date = f"{year}-01-01"
        end_date = cutoff if year == cutoff[:4] else f"{year}-12-31"
        data = api_get(
            f"{MLB_BASE}/schedule",
            {"sportId": 1, "venueIds": int(venue_id), "startDate": start_date, "endDate": end_date, "hydrate": "linescore"},
        )
        for day in data.get("dates", []):
            for game in day.get("games", []):
                if game.get("gameType") != "R":
                    continue
                if game.get("status", {}).get("codedGameState") != "F":
                    continue
                if game.get("officialDate") and game["officialDate"] >= cutoff:
                    continue
                away_score = stat_float(game.get("teams", {}).get("away", {}).get("score", 0))
                home_score = stat_float(game.get("teams", {}).get("home", {}).get("score", 0))
                away_abbr = game.get("teams", {}).get("away", {}).get("team", {}).get("name", "UNK")[:3].upper()
                home_abbr = game.get("teams", {}).get("home", {}).get("team", {}).get("name", "UNK")[:3].upper()
                date_fmt = game.get("officialDate", "")[5:].replace("-", "/")
                desc = f"{date_fmt} {away_abbr} {int(away_score)} @ {home_abbr} {int(home_score)}"
                games.append({
                    "date": game.get("officialDate"),
                    "desc": desc,
                    "home_runs": home_score,
                    "away_runs": away_score,
                    "total_runs": away_score + home_score,
                })
    games.sort(key=lambda x: x["date"], reverse=True)
    games = games[:limit]
    count = len(games)
    return {
        "games": count,
        "home_runs_avg": round(safe_div(sum(g["home_runs"] for g in games), count, 4.45), 2),
        "away_runs_avg": round(safe_div(sum(g["away_runs"] for g in games), count, 4.45), 2),
        "total_runs_avg": round(safe_div(sum(g["total_runs"] for g in games), count, 8.9), 2),
        "dates": [g.get("desc", g["date"]) for g in games],
    }

def fetch_schedule(game_date):
    params = {"sportId": 1, "date": game_date, "hydrate": "probablePitcher,team,venue"}
    data = api_get(f"{MLB_BASE}/schedule", params)
    games = []
    for d in data.get("dates", []):
        for g in d.get("games", []):
            games.append(g)
    return games

def get_team_stats(team_id):
    hitting = api_get(f"{MLB_BASE}/teams/{team_id}/stats", {"stats": "season", "group": "hitting"}).get("stats", [])
    pitching = api_get(f"{MLB_BASE}/teams/{team_id}/stats", {"stats": "season", "group": "pitching"}).get("stats", [])
    h = (hitting[0].get("splits", [{}])[0].get("stat", {}) if hitting else {})
    p = (pitching[0].get("splits", [{}])[0].get("stat", {}) if pitching else {})
    h_games = float(h.get("gamesPlayed", 0) or 0)
    p_games = float(p.get("gamesPlayed", 0) or 0)
    return {
        "hitting": {
            "runs_per_game": safe_div(float(h.get("runs", 0) or 0), h_games, 4.45),
            "ops": float(h.get("ops", 0) or 0),
            "avg": float(h.get("avg", 0) or 0),
            "obp": float(h.get("obp", 0) or 0),
            "slg": float(h.get("slg", 0) or 0),
            "homeRuns": float(h.get("homeRuns", 0) or 0),
        },
        "pitching": {
            "era": float(p.get("era", 4.20) or 4.20),
            "whip": float(p.get("whip", 1.30) or 1.30),
            "k9": float(p.get("strikeoutsPer9Inn", 8.5) or 8.5),
        }
    }

def get_player_stats(person_id, group):
    data = api_get(f"{MLB_BASE}/people/{person_id}/stats", {"stats": "season", "group": group}).get("stats", [])
    return (data[0].get("splits", [{}])[0].get("stat", {}) if data else {})

def get_roster(team_id):
    data = api_get(f"{MLB_BASE}/teams/{team_id}/roster", {"rosterType": "active"})
    return data.get("roster", [])

def enrich_hitters(team_id, game_date, opp_pitcher_id=None, limit=9):
    roster = get_roster(team_id)
    players = []
    for item in roster:
        person = item.get("person", {})
        pos = item.get("position", {}).get("abbreviation", "")
        if pos == "P":
            continue
        pid = person.get("id")
        if not pid:
            continue
        stat = get_player_stats(pid, "hitting")
        pa = float(stat.get("plateAppearances", 0) or 0)
        ops = float(stat.get("ops", 0) or 0)
        avg = float(stat.get("avg", 0) or 0)
        obp = float(stat.get("obp", 0) or 0)
        slg = float(stat.get("slg", 0) or 0)
        hr = float(stat.get("homeRuns", 0) or 0)
        if pa == 0 and ops == 0:
            continue
        recent = summarize_recent_hitting(pid, game_date, limit=4)
        matchup = aggregate_vs_player_history(pid, "hitting", opp_pitcher_id) if opp_pitcher_id else {
            "games": 0, "pa": 0.0, "ab": 0, "hits": 0, "hr": 0, "strikeouts": 0, "avg": 0.0, "obp": 0.0, "slg": 0.0, "ops": 0.0, "era_allowed": None
        }
        players.append({
            "id": pid,
            "name": person.get("fullName", "Unknown"),
            "pos": pos,
            "pa": pa,
            "ops": ops,
            "avg": avg,
            "obp": obp,
            "slg": slg,
            "hr": hr,
            "recent": recent,
            "vs_pitcher": matchup,
        })
    players.sort(key=lambda x: (x["pa"], x["ops"]), reverse=True)
    return players[:limit]

def get_probable_pitcher(game_side, game_date):
    pp = (game_side.get("probablePitcher") or {})
    if not pp:
        return None
    person_id = pp.get("id")
    stat = get_player_stats(person_id, "pitching")
    recent = summarize_recent_pitching(person_id, game_date, limit=4)
    return {
        "id": person_id,
        "name": pp.get("fullName", "TBD"),
        "era": float(stat.get("era", 4.20) or 4.20),
        "whip": float(stat.get("whip", 1.30) or 1.30),
        "k9": float(stat.get("strikeoutsPer9Inn", 8.5) or 8.5),
        "ip_per_start": float(stat.get("inningsPitched", 0) or 0) / max(float(stat.get("gamesStarted", 1) or 1), 1.0),
        "handedness": stat.get("pitchHand", {}).get("code") if isinstance(stat.get("pitchHand"), dict) else None,
        "recent": recent,
    }

def parse_odds_trader_state(html):
    match = re.search(r"window\.__INITIAL_STATE__=(\{.*?\});", html)
    if not match:
        return {}
    return json.loads(match.group(1))

def parse_odds_trader_hourly_dt(hour):
    utc_date = str(hour.get("utcDate") or "")
    utc_time = int(hour.get("utcTime") or 0)
    hour_part = utc_time // 100
    minute_part = utc_time % 100
    return datetime.fromisoformat(f"{utc_date}T{hour_part:02d}:{minute_part:02d}:00+00:00")

def nearest_hourly_weather(hourly, target_dt):
    if not hourly:
        return {}, 0
    best = hourly[0]
    best_index = 0
    best_diff = abs((parse_odds_trader_hourly_dt(best) - target_dt).total_seconds())
    for index, entry in enumerate(hourly[1:], start=1):
        diff = abs((parse_odds_trader_hourly_dt(entry) - target_dt).total_seconds())
        if diff < best_diff:
            best = entry
            best_index = index
            best_diff = diff
    return best, best_index

def format_odds_trader_hour_label(value):
    hour_value = int(stat_float(value, 0))
    hour = hour_value // 100
    minute = hour_value % 100
    suffix = "AM" if hour < 12 else "PM"
    hour12 = hour % 12 or 12
    if minute:
        return f"{hour12}:{minute:02d}{suffix}"
    return f"{hour12}{suffix}"

@lru_cache(maxsize=8)
def fetch_odds_trader_weather_board(board_date):
    html = requests.get(
        ODDS_TRADER_WEATHER_URL,
        timeout=25,
        headers={"User-Agent": USER_AGENT},
    ).text
    state = parse_odds_trader_state(html)
    weather_v2 = state.get("weatherV2", {}) or {}
    events = weather_v2.get("events", []) or []
    weather_by_event = weather_v2.get("weatherByEvent", {}) or {}
    out = []
    for event in events:
        away = next((p for p in event.get("participants", []) if not p.get("ih")), {})
        home = next((p for p in event.get("participants", []) if p.get("ih")), {})
        hourly = (weather_by_event.get(str(event.get("eid"))) or {}).get("hourly", [])
        out.append({
            "eid": str(event.get("eid")),
            "away_abbr": odds_trader_abbr((away.get("source") or {}).get("abbr")),
            "home_abbr": odds_trader_abbr((home.get("source") or {}).get("abbr")),
            "venue": str(event.get("ven") or ""),
            "field_out_label": str(event.get("fo") or "").upper(),
            "field_bearing_deg": direction_to_degrees(event.get("fo")),
            "game_dt": datetime.fromtimestamp((event.get("dt") or 0) / 1000).astimezone(),
            "hourly": hourly,
        })
    return out

@lru_cache(maxsize=8)
def fetch_ballpark_pal_factors():
    response = requests.get(
        BALLPARK_PAL_FACTORS_URL,
        timeout=25,
        headers={"User-Agent": USER_AGENT},
    )
    response.raise_for_status()
    rows = response.json()
    out = {}
    for row in rows if isinstance(rows, list) else []:
        game_pk = int(stat_float(row.get("GamePk"), 0))
        if not game_pk:
            continue
        out[game_pk] = {
            "game_pk": game_pk,
            "runs": stat_float(row.get("Runs"), 0.0),
            "hr": stat_float(row.get("HomeRuns"), 0.0),
            "doubles_triples": stat_float(row.get("DoublesTriples"), 0.0),
            "singles": stat_float(row.get("Singles"), 0.0),
            "runs_formatted": row.get("RunsFormatted", "0%"),
            "hr_formatted": row.get("HomeRunsFormatted", "0%"),
            "doubles_triples_formatted": row.get("DoublesTriplesFormatted", "0%"),
            "singles_formatted": row.get("SinglesFormatted", "0%"),
            "venue": row.get("VenueNameFull") or row.get("VenueName"),
            "game_time": row.get("GameTime"),
            "source": "ballparkpal",
        }
    return out

def get_ballpark_pal_factors(game_pk):
    return fetch_ballpark_pal_factors().get(int(game_pk), default_park_factors())

def find_odds_trader_event(game_date, venue, away_abbr, home_abbr):
    target_date = str(game_date)[:10]
    target_away = odds_trader_abbr(away_abbr)
    target_home = odds_trader_abbr(home_abbr)
    candidates = fetch_odds_trader_weather_board(target_date)
    exact = [
        event for event in candidates
        if event["away_abbr"] == target_away and event["home_abbr"] == target_home
    ]
    if len(exact) == 1:
        return exact[0]
    venue_lower = str(venue or "").strip().lower()
    venue_matches = [event for event in exact if event["venue"].strip().lower() == venue_lower]
    if venue_matches:
        return venue_matches[0]
    date_matches = [
        event for event in candidates
        if event["venue"].strip().lower() == venue_lower and str(event["game_dt"])[:10] == target_date
    ]
    return date_matches[0] if date_matches else None

def get_weather(game_date, venue, away_abbr, home_abbr):
    event = find_odds_trader_event(game_date, venue, away_abbr, home_abbr)
    if not event:
        return {
            "temperature_f": 70.0,
            "feels_like_f": 70.0,
            "cloud_cover": 50,
            "humidity": None,
            "precip_probability": 0,
            "wind_mph": 0.0,
            "wind_direction_deg": 0.0,
            "surface_pressure_hpa": None,
            "time_local": str(game_date),
            "weather_description": "Weather unavailable",
            "source": "oddstrader-unavailable",
            "field_bearing_deg": None,
            "field_out_label": None,
            "hourly_forecast": [],
        }

    target_dt = datetime.fromisoformat(str(game_date).replace("Z", "+00:00"))
    nearest, nearest_index = nearest_hourly_weather(event["hourly"], target_dt)
    hourly_forecast = []
    for entry in event["hourly"][nearest_index:nearest_index + 4]:
        hourly_forecast.append({
            "label": format_odds_trader_hour_label(entry.get("time")),
            "temperature_f": round(stat_float(entry.get("temperature"), 70.0), 1),
            "precip_probability": round(stat_float(entry.get("chanceOfRain"), entry.get("precipitation", 0.0)), 1),
            "wind_mph": round(stat_float(entry.get("windSpeedInMiles"), 0.0), 1),
            "wind_label": str(entry.get("windCardinal") or wind_label(entry.get("windDegree"))),
            "description": entry.get("weatherDescription", "Unknown"),
        })
    return {
        "temperature_f": round(stat_float(nearest.get("temperature"), 70.0), 1),
        "feels_like_f": round(stat_float(nearest.get("feelsLikeF"), nearest.get("temperature", 70.0)), 1),
        "cloud_cover": round(stat_float(nearest.get("cloudCover"), 50.0), 1),
        "humidity": None,
        "precip_probability": round(stat_float(nearest.get("chanceOfRain"), nearest.get("precipitation", 0.0)), 1),
        "wind_mph": round(stat_float(nearest.get("windSpeedInMiles"), 0.0), 1),
        "wind_direction_deg": normalize_degrees(nearest.get("windDegree")),
        "surface_pressure_hpa": None,
        "time_local": str(game_date),
        "weather_description": nearest.get("weatherDescription", "Unknown"),
        "source": "oddstrader",
        "field_bearing_deg": event.get("field_bearing_deg"),
        "field_out_label": event.get("field_out_label"),
        "hourly_forecast": hourly_forecast,
    }

def wind_label(deg):
    directions = ["N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE", "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW"]
    idx = round((normalize_degrees(deg) / 22.5)) % 16
    return directions[idx]

@lru_cache(maxsize=256)
def get_venue_orientation(venue_id):
    if not venue_id:
        return {}
    try:
        venue = api_get(f"{MLB_BASE}/venues/{int(venue_id)}", {"hydrate": "location"}).get("venues", [{}])[0]
        location = venue.get("location", {}) or {}
        field_bearing = location.get("azimuthAngle")
        if field_bearing in (None, ""):
            return {}
        field_bearing = normalize_degrees(field_bearing)
        return {
            "field_bearing_deg": round(field_bearing, 1),
            "out_direction": wind_label(field_bearing),
            "orientation_source": "mlb_statsapi",
        }
    except Exception:
        return {}

def park_wind_context(weather, field_bearing_deg=None):
    wind_from_deg = normalize_degrees(weather.get("wind_direction_deg"))
    wind_to_deg = (wind_from_deg + 180) % 360
    if field_bearing_deg is None:
        field_bearing_deg = 0.0
    field_bearing_deg = normalize_degrees(field_bearing_deg)
    relative_deg = (wind_to_deg - field_bearing_deg) % 360
    signed_relative = signed_angle_diff(wind_to_deg, field_bearing_deg)
    out_component = math.cos(math.radians(signed_relative))
    side_component = math.sin(math.radians(signed_relative))

    vertical = ""
    if out_component >= 0.15:
        vertical = "OUT"
    elif out_component <= -0.15:
        vertical = "IN"

    horizontal = ""
    if side_component >= 0.25:
        horizontal = "RF"
    elif side_component <= -0.25:
        horizontal = "LF"

    if vertical and horizontal:
        short_label = f"{vertical}-{horizontal}"
        text_label = f"{vertical.title()} and toward {horizontal}"
    elif vertical:
        short_label = vertical
        text_label = vertical.title()
    elif horizontal:
        short_label = horizontal
        text_label = f"Toward {horizontal}"
    else:
        short_label = "CALM"
        text_label = "Neutral"

    return {
        "wind_from_label": wind_label(wind_from_deg),
        "wind_to_deg": round(wind_to_deg, 1),
        "wind_relative_deg": round(relative_deg, 1),
        "wind_field_label": short_label,
        "wind_field_text": text_label,
        "wind_out_component": out_component,
        "wind_side_component": side_component,
    }

def run_environment_factor(weather, altitude_ft, park_factor, weights, field_bearing_deg=None):
    temperature_f = stat_float(weather.get("temperature_f"), 70.0)
    humidity = stat_float(weather.get("humidity"), 50.0)
    precip_probability = stat_float(weather.get("precip_probability"), 0.0)
    wind_mph = stat_float(weather.get("wind_mph"), 0.0)

    temp_factor = 1 + ((temperature_f - 70) / 10) * weights["weather_temp_per_10f"]
    humidity_factor = 1 + ((humidity - 50) / 10) * weights["humidity_per_10pct"]
    precip_factor = 1 + (precip_probability / 100) * weights["precip_penalty"]
    altitude_factor = 1 + (altitude_ft / 1000) * weights["altitude_per_1000ft"]

    wind_context = park_wind_context(weather, field_bearing_deg)
    out_score = wind_context["wind_out_component"]
    if out_score >= 0:
        wind_factor = 1 + (wind_mph / 10) * out_score * weights["wind_out_per_10mph"]
    else:
        wind_factor = 1 - (wind_mph / 10) * abs(out_score) * abs(weights["wind_in_per_10mph"])

    return max(0.82, min(1.28, temp_factor * humidity_factor * precip_factor * altitude_factor * wind_factor * park_factor))

def blend_rate(season_value, recent_value, recent_games, recent_weight=0.22):
    if recent_games <= 0:
        return season_value
    return (season_value * (1 - recent_weight)) + (recent_value * recent_weight)

# ── Weather impact on pitcher performance ──────────────────────────────
def pitcher_weather_factors(weather, altitude_ft, weights, field_bearing_deg=None):
    """Return multipliers that adjust pitcher stats for game-day conditions."""
    temp_f = stat_float(weather.get("temperature_f"), 70.0)
    humidity = stat_float(weather.get("humidity"), 50.0)
    precip_prob = stat_float(weather.get("precip_probability"), 0.0)
    wind_mph = stat_float(weather.get("wind_mph"), 0.0)

    wind_ctx = park_wind_context(weather, field_bearing_deg)
    out_component = wind_ctx["wind_out_component"]

    # Stamina: heat + humidity drain pitcher endurance → fewer IP
    heat_drain = max(0.0, (temp_f - 78) / 10) * 0.035       # lose ~0.35 IP per 10°F above 78
    humidity_drain = max(0.0, (humidity - 55) / 10) * 0.018  # lose ~0.18 IP per 10% humidity above 55
    stamina_factor = max(0.88, 1.0 - heat_drain - humidity_drain)

    # K-rate: cold reduces grip/spin → fewer K's; heat helps spin but fatigue hurts
    cold_k_penalty = max(0.0, (55 - temp_f) / 10) * 0.028   # ~2.8% fewer K per 10°F below 55
    heat_k_bonus = max(0.0, (temp_f - 75) / 10) * 0.012     # slight spin benefit in warmth
    k_rate_factor = max(0.90, min(1.08, 1.0 + heat_k_bonus - cold_k_penalty))

    # ERA/runs: wind blowing out = more hits/HR allowed; altitude thins air
    wind_era_adj = 0.0
    if out_component > 0:
        wind_era_adj = (wind_mph / 10) * out_component * 0.025
    elif out_component < 0:
        wind_era_adj = -(wind_mph / 10) * abs(out_component) * 0.018
    altitude_era_adj = (altitude_ft / 1000) * 0.012
    precip_era_adj = -(precip_prob / 100) * 0.02  # rain/mist slightly suppresses offense
    era_factor = max(0.88, min(1.18, 1.0 + wind_era_adj + altitude_era_adj + precip_era_adj))

    return {
        "stamina_factor": round(stamina_factor, 4),
        "k_rate_factor": round(k_rate_factor, 4),
        "era_factor": round(era_factor, 4),
    }

# ── Bullpen projection after starter exits ─────────────────────────────
def project_bullpen(starter_proj_ip, team_pitching_stats, weather_pitcher_factors):
    """Project bullpen contribution for innings after the starter."""
    total_game_ip = 9.0
    bullpen_ip = max(0.0, total_game_ip - starter_proj_ip)
    if bullpen_ip <= 0:
        return {
            "bullpen_ip": 0.0, "bullpen_era": 0.0, "bullpen_k": 0.0,
            "bullpen_er": 0.0, "bullpen_whip": 0.0,
        }
    # Team bullpen stats: use team pitching as proxy, inflate ERA slightly
    # (bullpen often has higher ERA than starters in aggregate)
    team_era = stat_float(team_pitching_stats.get("era"), 4.20)
    team_whip = stat_float(team_pitching_stats.get("whip"), 1.30)
    team_k9 = stat_float(team_pitching_stats.get("k9"), 8.5)
    bullpen_era = team_era * 1.05 * weather_pitcher_factors["era_factor"]
    bullpen_whip = team_whip * 1.03
    bullpen_k9 = team_k9 * 0.97 * weather_pitcher_factors["k_rate_factor"]
    bullpen_er = round(max(0.0, bullpen_ip * bullpen_era / 9), 1)
    bullpen_k = round(max(0.0, bullpen_ip * bullpen_k9 / 9), 1)
    return {
        "bullpen_ip": round(bullpen_ip, 1),
        "bullpen_era": round(bullpen_era, 2),
        "bullpen_whip": round(bullpen_whip, 2),
        "bullpen_k9": round(bullpen_k9, 1),
        "bullpen_k": bullpen_k,
        "bullpen_er": bullpen_er,
    }

# ── Weather impact on batter performance ───────────────────────────────
def batter_weather_factors(weather, altitude_ft, weights, field_bearing_deg=None):
    """Return multipliers that adjust batter stats for game-day conditions."""
    temp_f = stat_float(weather.get("temperature_f"), 70.0)
    humidity = stat_float(weather.get("humidity"), 50.0)
    precip_prob = stat_float(weather.get("precip_probability"), 0.0)
    wind_mph = stat_float(weather.get("wind_mph"), 0.0)

    wind_ctx = park_wind_context(weather, field_bearing_deg)
    out_component = wind_ctx["wind_out_component"]

    # Contact / AVG: warm temps = livelier ball; cold = harder to barrel
    temp_contact = 1 + ((temp_f - 70) / 10) * 0.008           # ~0.8% per 10°F
    precip_contact = 1 - (precip_prob / 100) * 0.025           # wet = slippery bat/ball
    contact_factor = max(0.92, min(1.10, temp_contact * precip_contact))

    # OBP: walks increase in bad weather (pitcher control suffers)
    walk_boost = (precip_prob / 100) * 0.015 + max(0.0, (humidity - 60) / 10) * 0.006
    cold_walk = max(0.0, (50 - temp_f) / 10) * 0.01          # cold = less control
    obp_factor = max(0.95, min(1.08, 1.0 + walk_boost + cold_walk))

    # Power / SLG: wind out + heat + altitude all boost fly ball distance
    wind_power = 0.0
    if out_component > 0:
        wind_power = (wind_mph / 10) * out_component * 0.032
    elif out_component < 0:
        wind_power = -(wind_mph / 10) * abs(out_component) * 0.025
    temp_power = ((temp_f - 70) / 10) * 0.014
    alt_power = (altitude_ft / 1000) * 0.015
    power_factor = max(0.85, min(1.22, 1.0 + wind_power + temp_power + alt_power))

    # Strikeout susceptibility: cold = harder to see spin; heat = better reaction
    cold_k_increase = max(0.0, (58 - temp_f) / 10) * 0.02
    k_susceptibility = max(0.94, min(1.08, 1.0 + cold_k_increase))

    return {
        "contact_factor": round(contact_factor, 4),
        "obp_factor": round(obp_factor, 4),
        "power_factor": round(power_factor, 4),
        "k_susceptibility": round(k_susceptibility, 4),
    }

# ── Times-through-the-order progression ────────────────────────────────
def times_through_order_factor(batting_slot, starter_proj_ip):
    """
    Batters improve on 2nd/3rd look at the same pitcher.  Top-of-order hitters
    see the starter more times.  Returns a multiplier > 1 if the batter benefits
    from repeated exposure, < 1 if they mostly face the starter once.
    """
    # Estimate how many times this slot sees the starter
    # slots 1-3 see starter ~3 times, slots 7-9 see ~2 times on avg in 5-6 IP
    if starter_proj_ip >= 6.0:
        times_faced = max(1.5, 3.2 - (batting_slot - 1) * 0.18)
    elif starter_proj_ip >= 5.0:
        times_faced = max(1.3, 2.8 - (batting_slot - 1) * 0.16)
    else:
        times_faced = max(1.0, 2.3 - (batting_slot - 1) * 0.14)

    # 2nd time through: ~7% better; 3rd time through: ~12% better (MLB research)
    if times_faced >= 2.5:
        progression_bonus = 0.045   # weighted avg of 2nd + 3rd look benefits
    elif times_faced >= 1.8:
        progression_bonus = 0.028   # mostly 2nd look benefit
    else:
        progression_bonus = 0.008   # minimal benefit, mostly first look

    # Late-game factor: batters in slots 1-5 more likely to face bullpen
    # (which they haven't seen), slightly reducing the progression bonus
    bullpen_dilution = 0.0
    if starter_proj_ip < 6.0 and batting_slot <= 5:
        bullpen_dilution = 0.01  # facing fresh relievers offsets some progression

    return round(1.0 + progression_bonus - bullpen_dilution, 4)

def project_team_runs(offense, recent_team, opp_pitcher, park_recent, weather_factor, park_factor, park_factors, is_home, weights):
    """Legacy score — kept for backwards compatibility, used as one input to simulate_game_score."""
    base = offense["runs_per_game"] * weights["offense_rpg_weight"] + weights["league_runs_per_team"] * (1 - weights["offense_rpg_weight"])
    recent_runs = recent_team.get("runs_per_game", base)
    base = base * 0.82 + recent_runs * 0.18

    pitcher_era = opp_pitcher["era"] * 0.75 + (opp_pitcher.get("recent", {}).get("era", opp_pitcher["era"]) * 0.25)
    pitcher_whip = opp_pitcher["whip"] * 0.75 + (opp_pitcher.get("recent", {}).get("whip", opp_pitcher["whip"]) * 0.25)
    era_factor = 1 + ((pitcher_era - 4.00) * weights["pitcher_era_weight"])
    whip_factor = 1 + ((pitcher_whip - 1.25) * weights["pitcher_whip_weight"])

    venue_avg = park_recent["home_runs_avg"] if is_home else park_recent["away_runs_avg"]
    venue_factor = 1 + ((venue_avg - weights["league_runs_per_team"]) / weights["league_runs_per_team"]) * 0.18
    total_factor = 1 + ((park_recent["total_runs_avg"] - (weights["league_runs_per_team"] * 2)) / (weights["league_runs_per_team"] * 2)) * 0.12
    park_board_factor = max(0.86, min(1.18, 1 + (stat_float(park_factors.get("runs"), 0.0) * 0.45)))

    return round(max(2.0, min(10.8, base * era_factor * whip_factor * weather_factor * park_factor * venue_factor * total_factor * park_board_factor)), 1)


# ── Full-data game score simulation ────────────────────────────────────
def simulate_game_score(
    offense, recent_team, opp_pitcher, opp_pitcher_proj, opp_bullpen,
    hitter_projs, park_recent, weather_factor, park_factor, park_factors,
    pitcher_wx, batter_wx, is_home, weights,
):
    """
    Composite score simulation using every available signal, weighted by
    predictive value from standard sabermetric research.

    Signals and approximate weights:
      Starting pitcher quality   28%   (ERA, WHIP, K/9, stamina, weather)
      Team offense (season)      22%   (RPG, OPS blended with league avg)
      Bullpen quality            13%   (remaining IP × bullpen ERA)
      Lineup vs pitcher matchup  10%   (individual batter-vs-pitcher OPS)
      Weather / environment       8%   (temp, wind, humidity, precip, altitude)
      Park factors                7%   (BallparkPal + venue recent scoring)
      Recent team form            6%   (last 4 games runs scored / allowed)
      Times-through-the-order     3%   (lineup progression benefit)
      Home-field advantage        3%   (historical ~54% home win rate)
    """
    league_rpg = weights.get("league_runs_per_team", 4.45)

    # ─── 1. Starting pitcher signal (28%) ──────────────────────────────
    # How many runs does this starter suppress / allow vs league average?
    if opp_pitcher and opp_pitcher.get("era") is not None:
        p_era = opp_pitcher["era"]
        p_whip = opp_pitcher.get("whip", 1.30)
        p_k9 = opp_pitcher.get("k9", 8.5)
        p_recent = opp_pitcher.get("recent", {})
        blended_era = p_era * 0.72 + p_recent.get("era", p_era) * 0.28
        blended_whip = p_whip * 0.72 + p_recent.get("whip", p_whip) * 0.28
        blended_k9 = p_k9 * 0.72 + p_recent.get("k9", p_k9) * 0.28
        # Starter projected IP (weather-adjusted)
        starter_ip = opp_pitcher_proj.get("proj_ip", 5.5)
        # ERA → runs/9 → runs in projected IP, adjusted by weather
        starter_runs_per_9 = blended_era * pitcher_wx.get("era_factor", 1.0)
        # WHIP penalty: high WHIP = more baserunners = more runs
        whip_adj = 1 + (blended_whip - 1.25) * weights.get("pitcher_whip_weight", 0.08)
        # K/9 suppression: high K rate = fewer balls in play = fewer runs
        k9_adj = 1 - (blended_k9 - 8.5) / 9 * 0.04
        starter_signal = (starter_runs_per_9 / 9) * starter_ip * whip_adj * k9_adj
    else:
        starter_ip = 5.0
        starter_signal = league_rpg * (starter_ip / 9.0)

    # ─── 2. Team offense signal (22%) ──────────────────────────────────
    season_rpg = offense.get("runs_per_game", league_rpg)
    season_ops = offense.get("ops", 0.720)
    # OPS-based run estimator: each .001 OPS ≈ ~0.005 RPG above/below avg
    ops_runs = league_rpg + (season_ops - 0.720) * 5.0
    offense_signal = season_rpg * 0.65 + ops_runs * 0.35

    # ─── 3. Bullpen signal (13%) ───────────────────────────────────────
    bullpen_ip = opp_bullpen.get("bullpen_ip", 9.0 - starter_ip)
    bullpen_era = opp_bullpen.get("bullpen_era", 4.40)
    bullpen_signal = (bullpen_era / 9) * bullpen_ip

    # ─── 4. Lineup vs pitcher matchup signal (10%) ─────────────────────
    matchup_ops_vals = []
    for h in (hitter_projs or []):
        vs = h.get("vs_pitcher", {})
        if vs.get("pa", 0) >= 4:
            matchup_ops_vals.append(vs.get("ops", 0.720))
    if matchup_ops_vals:
        avg_matchup_ops = sum(matchup_ops_vals) / len(matchup_ops_vals)
        matchup_signal = league_rpg * (1 + (avg_matchup_ops - 0.720) * 1.2)
    else:
        matchup_signal = league_rpg  # neutral when no data

    # ─── 5. Weather / environment signal (8%) ──────────────────────────
    # env_factor already combines temp, humidity, precip, altitude, wind, park
    # Batter weather factors layer on top for contact/power
    batter_wx_composite = (
        batter_wx.get("contact_factor", 1.0) * 0.30
        + batter_wx.get("obp_factor", 1.0) * 0.20
        + batter_wx.get("power_factor", 1.0) * 0.35
        + (2.0 - batter_wx.get("k_susceptibility", 1.0)) * 0.15  # invert: lower K = more runs
    )
    weather_signal = league_rpg * weather_factor * batter_wx_composite

    # ─── 6. Park factors signal (7%) ───────────────────────────────────
    venue_avg = park_recent.get("home_runs_avg", league_rpg) if is_home else park_recent.get("away_runs_avg", league_rpg)
    venue_adj = venue_avg / league_rpg if league_rpg else 1.0
    bp_runs_factor = stat_float(park_factors.get("runs"), 0.0)
    bp_hr_factor = stat_float(park_factors.get("hr"), 0.0)
    park_board_adj = 1 + bp_runs_factor * 0.35 + bp_hr_factor * 0.15
    total_park_adj = 1 + ((park_recent.get("total_runs_avg", league_rpg * 2) - league_rpg * 2) / (league_rpg * 2)) * 0.12
    park_signal = league_rpg * venue_adj * park_board_adj * total_park_adj * park_factor

    # ─── 7. Recent team form signal (6%) ───────────────────────────────
    recent_rpg = recent_team.get("runs_per_game", league_rpg)
    recent_signal = recent_rpg * 0.75 + league_rpg * 0.25

    # ─── 8. Times-through-the-order signal (3%) ───────────────────────
    tto_factors = [h.get("tto_factor", 1.0) for h in (hitter_projs or [])]
    avg_tto = sum(tto_factors) / len(tto_factors) if tto_factors else 1.0
    tto_signal = league_rpg * avg_tto

    # ─── 9. Home-field advantage signal (3%) ───────────────────────────
    home_edge = 1.024 if is_home else 0.976  # ~54% home win rate ≈ ±2.4% run boost

    # ─── Weighted composite ────────────────────────────────────────────
    # Pitching signals are "runs allowed TO this team" so they represent
    # what we expect the offense to score against that pitching.
    # Combine starter + bullpen into one pitching-allowed signal.
    pitching_allowed = starter_signal + bullpen_signal

    raw = (
        pitching_allowed * 0.28     # starter quality
        + offense_signal * 0.22     # team offense
        + pitching_allowed * 0.13   # bullpen (already in pitching_allowed, re-weight)
        + matchup_signal * 0.10     # lineup vs pitcher
        + weather_signal * 0.08     # weather / environment
        + park_signal * 0.07        # park factors
        + recent_signal * 0.06      # recent team form
        + tto_signal * 0.03         # times through order
        + league_rpg * 0.03         # home field (base to apply edge to)
    )

    # Apply home-field edge
    raw *= home_edge

    # Sanity clamp
    return round(max(1.5, min(12.0, raw)), 1)

def project_hitter_lines(hitters, team_runs, weather_factor, park_factors, batter_wx=None, opp_starter_proj_ip=5.5):
    default_bwx = {"contact_factor": 1.0, "obp_factor": 1.0, "power_factor": 1.0, "k_susceptibility": 1.0}
    bwx = batter_wx or default_bwx
    proj = []
    singles_boost = stat_float(park_factors.get("singles"), 0.0)
    doubles_triples_boost = stat_float(park_factors.get("doubles_triples"), 0.0)
    hr_boost = stat_float(park_factors.get("hr"), 0.0)
    runs_boost = stat_float(park_factors.get("runs"), 0.0)
    for i, h in enumerate(hitters[:9], start=1):
        slot_pa = 4.9 - (i - 1) * 0.12
        pa = round(max(3.3, slot_pa + (team_runs - 4.5) * 0.08), 1)
        recent = h.get("recent", {})
        matchup = h.get("vs_pitcher", {})
        form_avg = blend_rate(h["avg"], recent.get("avg", h["avg"]), recent.get("games", 0), recent_weight=0.24)
        form_obp = blend_rate(h["obp"], recent.get("obp", h["obp"]), recent.get("games", 0), recent_weight=0.24)
        form_slg = blend_rate(h["slg"], recent.get("slg", h["slg"]), recent.get("games", 0), recent_weight=0.24)
        form_ops = blend_rate(h["ops"], recent.get("ops", h["ops"]), recent.get("games", 0), recent_weight=0.24)

        matchup_factor = 1.0
        if matchup.get("pa", 0) >= 4:
            matchup_factor += max(-0.12, min(0.12, (matchup.get("ops", form_ops) - form_ops) * 0.45))

        # Times-through-the-order: batters improve on 2nd/3rd look at starter
        tto_factor = times_through_order_factor(i, opp_starter_proj_ip)

        # Park factors (unchanged)
        contact_park = max(0.88, min(1.18, 1 + (singles_boost * 0.45) + (doubles_triples_boost * 0.22)))
        power_park = max(0.82, min(1.24, 1 + (hr_boost * 0.7)))
        scoring_park = max(0.88, min(1.18, 1 + (runs_boost * 0.35)))

        # Combine park + weather for each stat category
        total_contact = contact_park * bwx["contact_factor"]
        total_obp_adj = max(0.9, min(1.14, 1 + (singles_boost * 0.25))) * bwx["obp_factor"]
        total_power = power_park * bwx["power_factor"]

        hits = max(0.4, min(2.4, pa * form_avg * matchup_factor * total_contact * tto_factor))
        ob = max(0.6, min(3.0, pa * form_obp * matchup_factor * total_obp_adj * tto_factor))
        hr_base = blend_rate(safe_div(h["hr"], max(h["pa"], 1), 0.0), safe_div(recent.get("hr", 0), max(recent.get("pa", 1), 1), 0.0), recent.get("games", 0), recent_weight=0.28)
        hr_matchup_boost = 1 + min(0.18, safe_div(matchup.get("hr", 0), max(matchup.get("pa", 1), 1), 0.0) * 2.0) if matchup.get("pa", 0) >= 6 else 1.0
        hr_rate = min(0.18, hr_base * 1.35 * weather_factor * hr_matchup_boost * total_power * tto_factor)
        rbi = max(0.2, min(2.5, team_runs * (0.06 + form_slg * 0.08) * matchup_factor * scoring_park * tto_factor))
        runs = max(0.2, min(2.2, team_runs * (0.05 + form_obp * 0.08) * matchup_factor * scoring_park * tto_factor))
        proj.append({
            "name": h["name"], "pos": h["pos"], "proj_pa": pa,
            "proj_hits": round(hits, 1), "proj_times_on_base": round(ob, 1),
            "proj_hr": round(min(0.9, hr_rate * pa), 2),
            "proj_rbi": round(rbi, 1), "proj_runs": round(runs, 1),
            "ops": round(form_ops, 3),
            "recent_form": recent,
            "vs_pitcher": matchup,
            "tto_factor": tto_factor,
        })
    return proj

def project_pitcher_line(pitcher, opp_runs, opp_hitters, weather_pitcher_fx=None, team_pitching_stats=None):
    default_wx = {"stamina_factor": 1.0, "k_rate_factor": 1.0, "era_factor": 1.0}
    wx = weather_pitcher_fx or default_wx
    if not pitcher:
        base_ip = 5.0
        bullpen = project_bullpen(base_ip, team_pitching_stats or {}, wx)
        return {
            "name": "TBD", "era": None, "whip": None, "k9": None,
            "proj_ip": base_ip, "proj_k": 4.0, "proj_er": round(opp_runs * 0.62 * wx["era_factor"], 1),
            "bullpen": bullpen,
            "weather_adjustments": wx,
        }
    recent = pitcher.get("recent", {})

    # IP: blend season/recent, then adjust for weather stamina
    raw_ip = (pitcher["ip_per_start"] * 0.74) + (recent.get("ip_per_start", pitcher["ip_per_start"]) * 0.26)
    ip = max(4.2, min(6.8, raw_ip * wx["stamina_factor"]))

    # K/9: blend season/recent, then adjust for weather grip/spin
    k9 = ((pitcher["k9"] * 0.76) + (recent.get("k9", pitcher["k9"]) * 0.24)) * wx["k_rate_factor"]

    # Matchup factor from opposing hitter history
    matchup_ops_values = [h.get("vs_pitcher", {}).get("ops", 0.0) for h in opp_hitters if h.get("vs_pitcher", {}).get("pa", 0) >= 3]
    matchup_factor = 1 + max(-0.08, min(0.08, ((sum(matchup_ops_values) / len(matchup_ops_values)) - 0.720) * 0.6)) if matchup_ops_values else 1.0

    k = max(2.5, min(10.5, ip * (k9 / 9) / matchup_factor))
    er = max(1.0, min(5.8, opp_runs * 0.62 * matchup_factor * wx["era_factor"]))

    # Bullpen projection for remaining innings
    bullpen = project_bullpen(ip, team_pitching_stats or {}, wx)

    return {
        "name": pitcher["name"], "era": pitcher["era"], "whip": pitcher["whip"], "k9": pitcher["k9"],
        "proj_ip": round(ip, 1), "proj_k": round(k, 1), "proj_er": round(er, 1),
        "recent": recent,
        "lineup_matchup": {
            "sampled_hitters": len(matchup_ops_values),
            "ops_avg": round(sum(matchup_ops_values) / len(matchup_ops_values), 3) if matchup_ops_values else None,
        },
        "bullpen": bullpen,
        "weather_adjustments": wx,
    }

@app.route("/")
def home():
    return render_template("index.html", today=date.today().isoformat())

@app.route("/api/games")
def games():
    game_date = request.args.get("date", date.today().isoformat())
    config = load_config()
    out = []
    for g in fetch_schedule(game_date):
        away = g["teams"]["away"]["team"]
        home = g["teams"]["home"]["team"]
        away_id, home_id = away["id"], home["id"]
        away_meta = config["parks"].get(int(away_id), {})
        home_meta = config["parks"].get(int(home_id), {})
        out.append({
            "gamePk": g["gamePk"],
            "gameDate": g.get("gameDate"),
            "status": g.get("status", {}).get("detailedState"),
            "venue": g.get("venue", {}).get("name", home_meta.get("venue")),
            "away": {"id": away_id, "name": away["name"], "abbr": away_meta.get("abbr", away["name"][:3].upper()), "logo": away_meta.get("logo_path")},
            "home": {"id": home_id, "name": home["name"], "abbr": home_meta.get("abbr", home["name"][:3].upper()), "logo": home_meta.get("logo_path")},
            "awayProbable": (g["teams"]["away"].get("probablePitcher") or {}).get("fullName", "TBD"),
            "homeProbable": (g["teams"]["home"].get("probablePitcher") or {}).get("fullName", "TBD"),
        })
    return jsonify({"date": game_date, "games": out})

@app.route("/api/projection/<int:game_pk>")
def projection(game_pk):
    config = load_config()
    games = fetch_schedule(request.args.get("date", date.today().isoformat()))
    game = next((x for x in games if int(x["gamePk"]) == int(game_pk)), None)
    if not game:
        return jsonify({"error": "Game not found for selected date"}), 404

    away = game["teams"]["away"]["team"]
    home = game["teams"]["home"]["team"]
    away_id, home_id = int(away["id"]), int(home["id"])
    venue_id = int((game.get("venue") or {}).get("id") or 0)
    home_meta = config["parks"][home_id]
    away_meta = config["parks"][away_id]
    weights = config["weights"]
    venue_orientation = get_venue_orientation(venue_id) if venue_id else {}
    weather = get_weather(game["gameDate"], home_meta["venue"], away_meta["abbr"], home_meta["abbr"])
    park_meta = {
        **home_meta,
        **venue_orientation,
        "field_bearing_deg": weather.get("field_bearing_deg", venue_orientation.get("field_bearing_deg")),
        "out_direction": weather.get("field_out_label") or venue_orientation.get("out_direction"),
    }

    wind_context = park_wind_context(weather, park_meta.get("field_bearing_deg"))
    field_bearing = park_meta.get("field_bearing_deg")
    env_factor = run_environment_factor(weather, home_meta["altitude_ft"], home_meta["park_factor"], weights, field_bearing)

    # Weather-specific factors for pitchers and batters
    pitcher_wx = pitcher_weather_factors(weather, home_meta["altitude_ft"], weights, field_bearing)
    batter_wx = batter_weather_factors(weather, home_meta["altitude_ft"], weights, field_bearing)

    away_team_stats = get_team_stats(away_id)
    home_team_stats = get_team_stats(home_id)
    away_recent = summarize_recent_team_runs(away_id, game["gameDate"], limit=4)
    home_recent = summarize_recent_team_runs(home_id, game["gameDate"], limit=4)
    park_recent = summarize_park_recent_scoring(venue_id, game["gameDate"], limit=8) if venue_id else {"games": 0, "home_runs_avg": 4.45, "away_runs_avg": 4.45, "total_runs_avg": 8.9, "dates": []}
    park_factors = get_ballpark_pal_factors(game_pk)

    away_pp = get_probable_pitcher(game["teams"]["away"], game["gameDate"])
    home_pp = get_probable_pitcher(game["teams"]["home"], game["gameDate"])

    away_pitcher = home_pp or {"era": 4.2, "whip": 1.30, "recent": {"era": 4.2, "whip": 1.30}}
    home_pitcher = away_pp or {"era": 4.2, "whip": 1.30, "recent": {"era": 4.2, "whip": 1.30}}

    # Legacy score (used as initial estimate for hitter RBI/runs projections)
    away_runs_legacy = project_team_runs(away_team_stats["hitting"], away_recent, away_pitcher, park_recent, env_factor, home_meta["park_factor"], park_factors, False, weights)
    home_runs_legacy = project_team_runs(home_team_stats["hitting"], home_recent, home_pitcher, park_recent, env_factor, home_meta["park_factor"], park_factors, True, weights)

    # Initial pitcher projections (before hitter matchup data)
    away_pitcher_proj = project_pitcher_line(away_pp, home_runs_legacy, [], pitcher_wx, away_team_stats["pitching"])
    home_pitcher_proj = project_pitcher_line(home_pp, away_runs_legacy, [], pitcher_wx, home_team_stats["pitching"])

    # Projected starter IP for times-through-the-order
    away_starter_ip = away_pitcher_proj["proj_ip"]
    home_starter_ip = home_pitcher_proj["proj_ip"]

    # Enrich hitters with matchup data
    away_enriched = enrich_hitters(away_id, game["gameDate"], opp_pitcher_id=home_pp["id"] if home_pp else None, limit=9)
    home_enriched = enrich_hitters(home_id, game["gameDate"], opp_pitcher_id=away_pp["id"] if away_pp else None, limit=9)

    away_hitters = project_hitter_lines(away_enriched, away_runs_legacy, env_factor, park_factors, batter_wx, home_starter_ip)
    home_hitters = project_hitter_lines(home_enriched, home_runs_legacy, env_factor, park_factors, batter_wx, away_starter_ip)

    # Re-project pitcher lines with actual hitter matchup data
    away_pitcher_proj = project_pitcher_line(away_pp, home_runs_legacy, home_hitters, pitcher_wx, away_team_stats["pitching"])
    home_pitcher_proj = project_pitcher_line(home_pp, away_runs_legacy, away_hitters, pitcher_wx, home_team_stats["pitching"])

    # ── Full-data simulated score ──────────────────────────────────────
    # Away team scores against home pitching (home_pp starter + home bullpen)
    away_sim = simulate_game_score(
        offense=away_team_stats["hitting"],
        recent_team=away_recent,
        opp_pitcher=home_pp or {"era": 4.2, "whip": 1.30, "k9": 8.5, "recent": {}},
        opp_pitcher_proj=home_pitcher_proj,
        opp_bullpen=home_pitcher_proj.get("bullpen", {}),
        hitter_projs=away_hitters,
        park_recent=park_recent,
        weather_factor=env_factor,
        park_factor=home_meta["park_factor"],
        park_factors=park_factors,
        pitcher_wx=pitcher_wx,
        batter_wx=batter_wx,
        is_home=False,
        weights=weights,
    )
    # Home team scores against away pitching (away_pp starter + away bullpen)
    home_sim = simulate_game_score(
        offense=home_team_stats["hitting"],
        recent_team=home_recent,
        opp_pitcher=away_pp or {"era": 4.2, "whip": 1.30, "k9": 8.5, "recent": {}},
        opp_pitcher_proj=away_pitcher_proj,
        opp_bullpen=away_pitcher_proj.get("bullpen", {}),
        hitter_projs=home_hitters,
        park_recent=park_recent,
        weather_factor=env_factor,
        park_factor=home_meta["park_factor"],
        park_factors=park_factors,
        pitcher_wx=pitcher_wx,
        batter_wx=batter_wx,
        is_home=True,
        weights=weights,
    )

    result = {
        "gamePk": game_pk,
        "gameDate": game["gameDate"],
        "status": game.get("status", {}).get("detailedState"),
        "venue": home_meta["venue"],
        "park": park_meta,
        "park_factors": park_factors,
        "park_recent_scoring": park_recent,
        "weather": {
            **weather,
            "wind_label": wind_label(weather["wind_direction_deg"]),
            **wind_context,
        },
        "environment_factor": round(env_factor, 3),
        "pitcher_weather_factors": pitcher_wx,
        "batter_weather_factors": batter_wx,
        "projected_score": {
            "away": away_sim,
            "home": home_sim,
            "total": round(away_sim + home_sim, 1),
        },
        "teams": {
            "away": {
                "id": away_id, "name": away["name"], "abbr": away_meta["abbr"], "logo": away_meta["logo_path"],
                "probable_pitcher": away_pitcher_proj,
                "offense_summary": away_team_stats["hitting"],
                "recent_team_form": away_recent,
                "hitters": away_hitters
            },
            "home": {
                "id": home_id, "name": home["name"], "abbr": home_meta["abbr"], "logo": home_meta["logo_path"],
                "probable_pitcher": home_pitcher_proj,
                "offense_summary": home_team_stats["hitting"],
                "recent_team_form": home_recent,
                "hitters": home_hitters
            }
        }
    }
    return jsonify(result)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(debug=True, host="0.0.0.0", port=port)
